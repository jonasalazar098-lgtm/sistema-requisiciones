from __future__ import annotations

from copy import copy
from io import BytesIO
from pathlib import Path
import os
import posixpath
import re
import shutil
import subprocess
import tempfile
import zipfile
import xml.etree.ElementTree as ET
from typing import Any

import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage, ImageDraw

try:
    import fitz  # PyMuPDF
except Exception:
    fitz = None


# =========================
# CONFIGURACIÓN
# =========================
st.set_page_config(
    page_title="Sistema de requisiciones",
    page_icon="📄",
    layout="wide",
)

BUILTIN_TEMPLATE_PATH = Path(__file__).with_name("FORMATOS.xlsx")


# =========================
# UTILIDADES
# =========================
def normalizar_texto(valor: Any) -> str:
    if valor is None:
        return ""
    return str(valor).strip()


def normalizar_mayus(valor: Any) -> str:
    return normalizar_texto(valor).upper()


def limpiar_numero(valor: Any) -> float | None:
    if valor is None:
        return None

    texto = str(valor)
    texto = texto.replace("$", "")
    texto = texto.replace(",", "")
    texto = texto.strip()

    if texto == "":
        return None

    try:
        return float(texto)
    except ValueError:
        return None


def es_entero(texto: str) -> bool:
    return bool(re.fullmatch(r"\d{1,8}", texto.strip()))


def es_numero(texto: str) -> bool:
    return bool(re.fullmatch(r"\$?\d+(?:,\d{3})*(?:\.\d+)?", texto.strip()))


def es_dinero(texto: str) -> bool:
    return bool(re.fullmatch(r"\$?\d+(?:,\d{3})*\.\d{2}", texto.strip()))


def elegir_precio_importe(cantidad: float | None, n1: float | None, n2: float | None):
    if cantidad is None or n1 is None or n2 is None:
        return n1, n2

    error_1 = abs((cantidad * n1) - n2)
    error_2 = abs((cantidad * n2) - n1)

    tolerancia_1 = max(0.10, abs(n2) * 0.001)
    tolerancia_2 = max(0.10, abs(n1) * 0.001)

    if error_1 <= tolerancia_1:
        return n1, n2

    if error_2 <= tolerancia_2:
        return n2, n1

    if n1 <= n2:
        return n1, n2

    return n2, n1


def nombre_hoja_valido(nombre: str) -> str:
    nombre = normalizar_texto(nombre) or "Requisicion"
    nombre = re.sub(r"[\[\]\:\*\?\/\\]", "_", nombre)
    return nombre[:31]


# =========================
# CELDAS COMBINADAS Y BÚSQUEDA
# =========================
def anchor_coord(ws, coord: str) -> str:
    """
    Si coord está dentro de una celda combinada, devuelve la celda ancla.
    """
    if not isinstance(ws[coord], MergedCell):
        return coord

    for rango in ws.merged_cells.ranges:
        if coord in rango:
            return ws.cell(rango.min_row, rango.min_col).coordinate

    return coord


def set_cell(ws, coord: str, value: Any):
    ws[anchor_coord(ws, coord)] = value


def get_cell(ws, coord: str) -> Any:
    return ws[anchor_coord(ws, coord)].value


def find_cell(ws, contains: str, start_row: int = 1, end_row: int | None = None):
    target = normalizar_mayus(contains)
    end_row = end_row or ws.max_row

    for row in range(start_row, end_row + 1):
        for col in range(1, ws.max_column + 1):
            value = normalizar_mayus(get_cell(ws, ws.cell(row, col).coordinate))
            if target in value:
                return row, col, value

    return None


def find_exact_cell(ws, text: str, start_row: int = 1, end_row: int | None = None):
    target = normalizar_mayus(text)
    end_row = end_row or ws.max_row

    for row in range(start_row, end_row + 1):
        for col in range(1, ws.max_column + 1):
            value = normalizar_mayus(get_cell(ws, ws.cell(row, col).coordinate))
            if value == target:
                return row, col, value

    return None


def write_right_of_label(ws, label: str, value: Any, offset_cols: int = 1):
    found = find_cell(ws, label)
    if found:
        row, col, _ = found
        set_cell(ws, ws.cell(row, col + offset_cols).coordinate, value)


def copy_row_style(ws, source_row: int, target_row: int, max_col: int):
    for col in range(1, max_col + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)

        if isinstance(target, MergedCell):
            continue

        if source.has_style:
            target.font = copy(source.font)
            target.border = copy(source.border)
            target.fill = copy(source.fill)
            target.number_format = source.number_format
            target.protection = copy(source.protection)
            target.alignment = copy(source.alignment)


def _reemplazar_borde(
    borde_actual: Border,
    left=None,
    right=None,
    top=None,
    bottom=None,
    diagonal=None,
    diagonal_up=None,
    diagonal_down=None,
) -> Border:
    return Border(
        left=left if left is not None else borde_actual.left,
        right=right if right is not None else borde_actual.right,
        top=top if top is not None else borde_actual.top,
        bottom=bottom if bottom is not None else borde_actual.bottom,
        diagonal=diagonal if diagonal is not None else borde_actual.diagonal,
        diagonalUp=diagonal_up if diagonal_up is not None else borde_actual.diagonalUp,
        diagonalDown=diagonal_down if diagonal_down is not None else borde_actual.diagonalDown,
        outline=borde_actual.outline,
        vertical=borde_actual.vertical,
        horizontal=borde_actual.horizontal,
    )


def _ancho_columna_px(ws, col: int) -> int:
    # No usar ws.cell(1, col).column_letter porque si la celda está combinada
    # puede ser MergedCell y provocar error. get_column_letter funciona siempre.
    letra = get_column_letter(col)
    ancho = ws.column_dimensions[letra].width
    if ancho is None:
        ancho = 8.43
    return max(12, int(float(ancho) * 7 + 5))


def _alto_fila_px(ws, row: int) -> int:
    alto = ws.row_dimensions[row].height
    if alto is None:
        alto = 15
    return max(10, int(float(alto) * 96 / 72))


def _crear_imagen_diagonal(ws, meta: dict[str, int]) -> str:
    """
    Crea una imagen PNG transparente con una diagonal gris.
    Se inserta sobre el rango vacío para que Excel y LibreOffice/PDF la respeten.
    """
    width = sum(_ancho_columna_px(ws, c) for c in range(meta["start_col"], meta["end_col"] + 1))
    height = sum(_alto_fila_px(ws, r) for r in range(meta["first_blank_row"], meta["last_blank_row"] + 1))

    width = max(width, 120)
    height = max(height, 40)

    img = PILImage.new("RGBA", (width, height), (255, 255, 255, 0))
    draw = ImageDraw.Draw(img)

    # Diagonal tipo cierre: de abajo-izquierda a arriba-derecha.
    draw.line((0, height - 1, width - 1, 0), fill=(90, 90, 90, 255), width=3)

    tmp_dir = Path(tempfile.gettempdir()) / "cierres_requisiciones"
    tmp_dir.mkdir(parents=True, exist_ok=True)
    path = tmp_dir / f"cierre_{os.getpid()}_{meta['first_blank_row']}_{meta['last_blank_row']}.png"
    img.save(path)
    return str(path)


def marcar_cierre_tabla(ws, primera_fila_vacia: int, ultima_fila_vacia: int, col_inicio: int, col_fin: int):
    """
    Cierre visual de la orden:
    - Franja gris inmediatamente debajo del último producto.
    - Una sola diagonal limpia sobre el espacio vacío restante.

    Se eliminan las diagonales pequeñas por celda para que el cierre se vea limpio.
    """
    gris_borde = Side(style="medium", color="404040")
    relleno_gris = PatternFill("solid", fgColor="A6A6A6")

    fila_ultimo_articulo = max(1, primera_fila_vacia - 1)
    fila_franja = primera_fila_vacia
    fila_inicio_diagonal = primera_fila_vacia + 1

    # 1) Borde/línea gris debajo del último producto.
    for col in range(col_inicio, col_fin + 1):
        celda_ultimo = ws.cell(fila_ultimo_articulo, col)
        if not isinstance(celda_ultimo, MergedCell):
            celda_ultimo.border = _reemplazar_borde(celda_ultimo.border, bottom=gris_borde)

    # 2) Franja gris visible justo debajo del último producto.
    if fila_franja <= ultima_fila_vacia:
        for col in range(col_inicio, col_fin + 1):
            celda = ws.cell(fila_franja, col)
            if not isinstance(celda, MergedCell):
                celda.fill = relleno_gris
                celda.value = None
                celda.border = _reemplazar_borde(
                    celda.border,
                    top=gris_borde,
                    bottom=gris_borde,
                )

    # Si no queda espacio después de la franja gris, no se dibuja diagonal.
    if fila_inicio_diagonal > ultima_fila_vacia:
        ws._cierre_visual = None
        return

    width_px = sum(_ancho_columna_px(ws, c) for c in range(col_inicio, col_fin + 1))
    height_px = sum(_alto_fila_px(ws, r) for r in range(fila_inicio_diagonal, ultima_fila_vacia + 1))

    # 3) Diagonal principal única como dibujo interno del XLSX.
    # El PDF se genera desde este mismo Excel.
    ws._cierre_visual = {
        "first_blank_row": fila_inicio_diagonal,
        "last_blank_row": ultima_fila_vacia,
        "start_col": col_inicio,
        "end_col": col_fin,
        "width_px": max(width_px, 120),
        "height_px": max(height_px, 40),
    }


def _normalizar_target(base_path: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join(posixpath.dirname(base_path), target))


def _siguiente_rid(rels_root) -> str:
    usados = []
    for rel in rels_root:
        rid = rel.attrib.get("Id", "")
        m = re.match(r"rId(\d+)$", rid)
        if m:
            usados.append(int(m.group(1)))
    return f"rId{(max(usados) + 1) if usados else 1}"


def _siguiente_drawing_path(nombres_zip: set[str]) -> str:
    usados = []
    for nombre in nombres_zip:
        m = re.match(r"xl/drawings/drawing(\d+)\.xml$", nombre)
        if m:
            usados.append(int(m.group(1)))
    n = (max(usados) + 1) if usados else 1
    return f"xl/drawings/drawing{n}.xml"


def _max_cnvpr_id(drawing_root) -> int:
    max_id = 1
    for elem in drawing_root.iter():
        if elem.tag.endswith("cNvPr"):
            try:
                max_id = max(max_id, int(elem.attrib.get("id", "1")))
            except ValueError:
                pass
    return max_id


def _crear_anchor_diagonal(meta: dict[str, int], shape_id: int):
    ns_xdr = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    ns_a = "http://schemas.openxmlformats.org/drawingml/2006/main"

    ancho_emu = str(int(meta.get("width_px", 800) * 9525))
    alto_emu = str(int(meta.get("height_px", 400) * 9525))

    anchor = ET.Element(f"{{{ns_xdr}}}twoCellAnchor", {"editAs": "twoCell"})

    desde = ET.SubElement(anchor, f"{{{ns_xdr}}}from")
    ET.SubElement(desde, f"{{{ns_xdr}}}col").text = str(meta["start_col"] - 1)
    ET.SubElement(desde, f"{{{ns_xdr}}}colOff").text = "0"
    ET.SubElement(desde, f"{{{ns_xdr}}}row").text = str(meta["first_blank_row"] - 1)
    ET.SubElement(desde, f"{{{ns_xdr}}}rowOff").text = "0"

    hasta = ET.SubElement(anchor, f"{{{ns_xdr}}}to")
    ET.SubElement(hasta, f"{{{ns_xdr}}}col").text = str(meta["end_col"])
    ET.SubElement(hasta, f"{{{ns_xdr}}}colOff").text = "0"
    ET.SubElement(hasta, f"{{{ns_xdr}}}row").text = str(meta["last_blank_row"])
    ET.SubElement(hasta, f"{{{ns_xdr}}}rowOff").text = "0"

    sp = ET.SubElement(anchor, f"{{{ns_xdr}}}sp", {"macro": "", "textlink": ""})
    nv = ET.SubElement(sp, f"{{{ns_xdr}}}nvSpPr")
    ET.SubElement(nv, f"{{{ns_xdr}}}cNvPr", {"id": str(shape_id), "name": "Cierre diagonal"})
    ET.SubElement(nv, f"{{{ns_xdr}}}cNvSpPr")

    sppr = ET.SubElement(sp, f"{{{ns_xdr}}}spPr")
    # flipV = línea de abajo-izquierda hacia arriba-derecha.
    xfrm = ET.SubElement(sppr, f"{{{ns_a}}}xfrm", {"flipV": "1"})
    ET.SubElement(xfrm, f"{{{ns_a}}}off", {"x": "0", "y": "0"})
    ET.SubElement(xfrm, f"{{{ns_a}}}ext", {"cx": ancho_emu, "cy": alto_emu})

    geom = ET.SubElement(sppr, f"{{{ns_a}}}prstGeom", {"prst": "line"})
    ET.SubElement(geom, f"{{{ns_a}}}avLst")

    # Línea gris más visible.
    ln = ET.SubElement(sppr, f"{{{ns_a}}}ln", {"w": "25400", "cap": "flat"})
    fill = ET.SubElement(ln, f"{{{ns_a}}}solidFill")
    ET.SubElement(fill, f"{{{ns_a}}}srgbClr", {"val": "404040"})

    ET.SubElement(anchor, f"{{{ns_xdr}}}clientData")
    return anchor


def agregar_cierre_diagonal_xlsx(excel_bytes: bytes, sheet_name: str, meta: dict[str, int]) -> bytes:
    """
    Inserta una línea diagonal real en la hoja generada modificando el XML interno del XLSX.
    Así el resultado se ve igual en Excel y en el PDF convertido por LibreOffice.
    """
    if not meta:
        return excel_bytes

    ET.register_namespace("r", "http://schemas.openxmlformats.org/officeDocument/2006/relationships")
    ET.register_namespace("xdr", "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing")
    ET.register_namespace("a", "http://schemas.openxmlformats.org/drawingml/2006/main")

    ns_main = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ns_rel = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    ns_pkg_rel = "http://schemas.openxmlformats.org/package/2006/relationships"
    ns_ct = "http://schemas.openxmlformats.org/package/2006/content-types"
    ns_xdr = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"

    with zipfile.ZipFile(BytesIO(excel_bytes), "r") as zin:
        archivos = {name: zin.read(name) for name in zin.namelist()}

    nombres = set(archivos.keys())

    workbook_root = ET.fromstring(archivos["xl/workbook.xml"])
    sheet_rid = None
    for sheet in workbook_root.findall(f".//{{{ns_main}}}sheet"):
        if sheet.attrib.get("name") == sheet_name:
            sheet_rid = sheet.attrib.get(f"{{{ns_rel}}}id")
            break

    if not sheet_rid:
        return excel_bytes

    wb_rels_root = ET.fromstring(archivos["xl/_rels/workbook.xml.rels"])
    sheet_path = None
    for rel in wb_rels_root:
        if rel.attrib.get("Id") == sheet_rid:
            target = rel.attrib.get("Target", "")
            sheet_path = target.lstrip("/") if target.startswith("/") else posixpath.normpath(posixpath.join("xl", target))
            break

    if not sheet_path or sheet_path not in archivos:
        return excel_bytes

    sheet_root = ET.fromstring(archivos[sheet_path])
    drawing_elem = sheet_root.find(f"{{{ns_main}}}drawing")

    rels_path = posixpath.join(posixpath.dirname(sheet_path), "_rels", posixpath.basename(sheet_path) + ".rels")
    if rels_path in archivos:
        sheet_rels_root = ET.fromstring(archivos[rels_path])
    else:
        sheet_rels_root = ET.Element("Relationships", {"xmlns": ns_pkg_rel})

    drawing_path = None
    if drawing_elem is not None:
        drawing_rid = drawing_elem.attrib.get(f"{{{ns_rel}}}id")
        for rel in sheet_rels_root:
            if rel.attrib.get("Id") == drawing_rid:
                drawing_path = _normalizar_target(sheet_path, rel.attrib.get("Target", ""))
                break

    if drawing_path and drawing_path in archivos:
        drawing_root = ET.fromstring(archivos[drawing_path])
    else:
        drawing_path = _siguiente_drawing_path(nombres)
        drawing_root = ET.Element(f"{{{ns_xdr}}}wsDr")

        nuevo_rid = _siguiente_rid(sheet_rels_root)
        ET.SubElement(
            sheet_rels_root,
            "Relationship",
            {
                "Id": nuevo_rid,
                "Type": "http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing",
                "Target": posixpath.relpath(drawing_path, posixpath.dirname(sheet_path)),
            },
        )
        drawing_elem = ET.Element(f"{{{ns_main}}}drawing", {f"{{{ns_rel}}}id": nuevo_rid})
        sheet_root.append(drawing_elem)

        # Agregar ContentType si es un dibujo nuevo.
        ct_root = ET.fromstring(archivos["[Content_Types].xml"])
        parte = "/" + drawing_path
        existe = any(elem.attrib.get("PartName") == parte for elem in ct_root.findall(f"{{{ns_ct}}}Override"))
        if not existe:
            ET.SubElement(
                ct_root,
                f"{{{ns_ct}}}Override",
                {
                    "PartName": parte,
                    "ContentType": "application/vnd.openxmlformats-officedocument.drawing+xml",
                },
            )
        archivos["[Content_Types].xml"] = ET.tostring(ct_root, encoding="utf-8", xml_declaration=True)

    shape_id = _max_cnvpr_id(drawing_root) + 1
    drawing_root.append(_crear_anchor_diagonal(meta, shape_id))

    archivos[drawing_path] = ET.tostring(drawing_root, encoding="utf-8", xml_declaration=True)
    archivos[rels_path] = ET.tostring(sheet_rels_root, encoding="utf-8", xml_declaration=True)
    archivos[sheet_path] = ET.tostring(sheet_root, encoding="utf-8", xml_declaration=True)

    salida = BytesIO()
    with zipfile.ZipFile(salida, "w", zipfile.ZIP_DEFLATED) as zout:
        for nombre, datos in archivos.items():
            zout.writestr(nombre, datos)

    return salida.getvalue()


# =========================
# LECTURA DE CAMPOS DE PLANTILLA
# =========================
def leer_derecha_de_label(ws, label: str, offset_cols: int = 1):
    found = find_cell(ws, label)
    if not found:
        return ""
    row, col, _ = found
    return get_cell(ws, ws.cell(row, col + offset_cols).coordinate) or ""


def leer_campos_requisicion(ws) -> dict[str, Any]:
    campos = {
        "folio": leer_derecha_de_label(ws, "No.", 1),
        "obra": leer_derecha_de_label(ws, "obra", 1),
        "proveedor": leer_derecha_de_label(ws, "proveedor", 1),
        "fecha": leer_derecha_de_label(ws, "fecha", 1),
        "semana": leer_derecha_de_label(ws, "semana", 1),
        "numero_obra": leer_derecha_de_label(ws, "No. Obra", 1),
        "id_familia": "",
        "etapa": "",
        "area": "",
        "partida": "",
        "uso": "",
        "solicita": "",
        "recibe": "",
        "entregar_en": "",
        "facturar_a": "",
        "firma_izquierda": "",
        "firma_derecha": "",
        "puesto_izquierdo": "",
        "puesto_derecho": "",
    }

    found = find_cell(ws, "ID FAMILIA")
    if found:
        row, col, _ = found
        campos["id_familia"] = get_cell(ws, ws.cell(row, min(col + 4, ws.max_column)).coordinate) or ""

    found = find_cell(ws, "ETAPA")
    if found:
        row, col, _ = found
        campos["etapa"] = get_cell(ws, ws.cell(row + 1, col).coordinate) or ""

    found = find_cell(ws, "ÁREA") or find_cell(ws, "AREA")
    if found:
        row, col, _ = found
        campos["area"] = get_cell(ws, ws.cell(row + 1, col).coordinate) or ""

    found = find_cell(ws, "PARTIDA")
    if found:
        row, col, _ = found
        campos["partida"] = get_cell(ws, ws.cell(row + 1, col).coordinate) or ""

    found = find_cell(ws, "OBSERVACIONES")
    if found:
        row, col, _ = found
        campos["uso"] = get_cell(ws, ws.cell(row + 1, col).coordinate) or ""

    found = find_cell(ws, "SOLICITA")
    if found:
        row, col, value = found
        campos["solicita"] = str(get_cell(ws, ws.cell(row, col).coordinate) or "").replace("SOLICITA:", "").strip()

    found = find_cell(ws, "RECIBE")
    if found:
        row, col, value = found
        campos["recibe"] = str(get_cell(ws, ws.cell(row, col).coordinate) or "").replace("RECIBE:", "").strip()

    found = find_cell(ws, "ENTREGAR EN")
    if found:
        row, col, _ = found
        campos["entregar_en"] = get_cell(ws, ws.cell(row + 1, col).coordinate) or ""

    found = find_cell(ws, "FACTURAR A")
    if found:
        row, col, _ = found
        campos["facturar_a"] = get_cell(ws, ws.cell(row, min(col + 2, ws.max_column)).coordinate) or ""

    # Firmas conocidas.
    for row in range(max(1, ws.max_row - 15), ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            txt = normalizar_mayus(get_cell(ws, ws.cell(row, col).coordinate))
            if "KARIME" in txt or "LIC." in txt:
                campos["firma_izquierda"] = get_cell(ws, ws.cell(row, col).coordinate) or campos["firma_izquierda"]
                campos["puesto_izquierdo"] = get_cell(ws, ws.cell(row + 1, col).coordinate) or campos["puesto_izquierdo"]
            if "JUAN ANGEL" in txt or "DIRECTOR" in txt:
                if "DIRECTOR" not in txt:
                    campos["firma_derecha"] = get_cell(ws, ws.cell(row, col).coordinate) or campos["firma_derecha"]
                    campos["puesto_derecho"] = get_cell(ws, ws.cell(row + 1, col).coordinate) or campos["puesto_derecho"]

    return campos


def leer_campos_orden_compra(ws) -> dict[str, Any]:
    campos = {
        "folio": "",
        "obra": "",
        "proveedor": leer_derecha_de_label(ws, "PROVEEDOR", 1),
        "fecha": leer_derecha_de_label(ws, "FECHA", 1),
        "semana": "",
        "numero_obra": "",
        "id_familia": "",
        "etapa": "",
        "area": "",
        "partida": "",
        "uso": leer_derecha_de_label(ws, "USO", 1),
        "solicita": leer_derecha_de_label(ws, "SOLICITA", 1),
        "recibe": "",
        "entregar_en": leer_derecha_de_label(ws, "ENTREGAR EN", 1),
        "facturar_a": leer_derecha_de_label(ws, "FACTURAR A", 1),
        "firma_izquierda": "",
        "firma_derecha": "",
        "puesto_izquierdo": "",
        "puesto_derecho": "",
    }

    found = find_cell(ws, "ORDEN DE COMPRA")
    if found:
        row, col, _ = found
        campos["folio"] = get_cell(ws, ws.cell(row + 1, col).coordinate) or ""

    return campos


# =========================
# DETECCIÓN DE TIPO DE PLANTILLA
# =========================
def detectar_tipo_plantilla(ws) -> str:
    if find_cell(ws, "ORDEN DE COMPRA"):
        return "orden_compra"
    return "requisicion"


def detectar_header_requisicion(ws):
    for row in range(1, min(ws.max_row, 20) + 1):
        textos = [normalizar_mayus(get_cell(ws, ws.cell(row, col).coordinate)) for col in range(1, ws.max_column + 1)]
        joined = " ".join(textos)
        if "PAR" in joined and "CANT" in joined and "DESCRIPCION" in joined:
            return row
    return 8


def detectar_columnas_requisicion(ws, header_row: int):
    cols = {
        "par": 1,
        "cantidad": 3,
        "unidad": 4,
        "descripcion": 5,
        "codigo": 8,
        "pu": 11,
        "total": 12,
    }

    for col in range(1, ws.max_column + 1):
        txt = normalizar_mayus(get_cell(ws, ws.cell(header_row, col).coordinate))
        if "PAR" == txt:
            cols["par"] = col
        elif "CANT" in txt:
            cols["cantidad"] = col
        elif "UNI" in txt:
            cols["unidad"] = col
        elif "DESCRIPCION" in txt or "DESCRIPCIÓN" in txt:
            cols["descripcion"] = col
        elif "CODIGO" in txt or "CÓDIGO" in txt:
            cols["codigo"] = col
        elif "P.U" in txt:
            cols["pu"] = col
        elif "TOTAL" == txt:
            cols["total"] = col

    return cols


def find_totals_row(ws, start_row: int):
    found = find_exact_cell(ws, "SUBTOTAL", start_row=start_row)
    if found:
        return found[0]
    found = find_cell(ws, "SUBTOTAL", start_row=start_row)
    if found:
        return found[0]
    return start_row + 40


def preparar_filas_tabla(ws, data_start: int, subtotal_row: int, partidas_count: int):
    capacidad = subtotal_row - data_start
    filas_extra = max(0, partidas_count - capacidad)

    if filas_extra:
        # Copiar siempre el formato de una fila normal de artículo.
        # Antes se copiaba la fila justo antes del subtotal y, en algunos formatos,
        # esa fila tenía alineación/estilo raro. Por eso los últimos productos
        # podían salir centrados o desacomodados cuando había muchas partidas.
        fila_estilo = data_start
        ws.insert_rows(subtotal_row, amount=filas_extra)
        for row in range(subtotal_row, subtotal_row + filas_extra):
            copy_row_style(ws, fila_estilo, row, ws.max_column)

    return filas_extra


# =========================
# LLENAR PLANTILLAS
# =========================
def aplicar_campos_requisicion(ws, campos: dict[str, Any]):
    write_right_of_label(ws, "No.", campos.get("folio", ""), 1)
    write_right_of_label(ws, "obra", campos.get("obra", ""), 1)
    write_right_of_label(ws, "proveedor", campos.get("proveedor", ""), 1)
    write_right_of_label(ws, "fecha", campos.get("fecha", ""), 1)
    write_right_of_label(ws, "semana", campos.get("semana", ""), 1)
    write_right_of_label(ws, "No. Obra", campos.get("numero_obra", ""), 1)

    found = find_cell(ws, "ID FAMILIA")
    if found:
        row, col, _ = found
        set_cell(ws, ws.cell(row, min(col + 4, ws.max_column)).coordinate, campos.get("id_familia", ""))

    found = find_cell(ws, "ETAPA")
    if found:
        row, col, _ = found
        set_cell(ws, ws.cell(row + 1, col).coordinate, campos.get("etapa", ""))

    found = find_cell(ws, "ÁREA") or find_cell(ws, "AREA")
    if found:
        row, col, _ = found
        set_cell(ws, ws.cell(row + 1, col).coordinate, campos.get("area", ""))

    found = find_cell(ws, "PARTIDA")
    if found:
        row, col, _ = found
        set_cell(ws, ws.cell(row + 1, col).coordinate, campos.get("partida", ""))

    found = find_cell(ws, "OBSERVACIONES")
    if found:
        row, col, _ = found
        set_cell(ws, ws.cell(row + 1, col).coordinate, campos.get("uso", ""))

    found = find_cell(ws, "SOLICITA")
    if found:
        row, col, _ = found
        set_cell(ws, ws.cell(row, col).coordinate, f"SOLICITA: {campos.get('solicita', '')}".strip())

    found = find_cell(ws, "RECIBE")
    if found:
        row, col, _ = found
        set_cell(ws, ws.cell(row, col).coordinate, f"RECIBE: {campos.get('recibe', '')}".strip())

    found = find_cell(ws, "ENTREGAR EN")
    if found:
        row, col, _ = found
        set_cell(ws, ws.cell(row + 1, col).coordinate, campos.get("entregar_en", ""))

    found = find_cell(ws, "FACTURAR A")
    if found:
        row, col, _ = found
        set_cell(ws, ws.cell(row, min(col + 2, ws.max_column)).coordinate, campos.get("facturar_a", ""))


def llenar_requisicion(ws, campos: dict[str, Any], partidas: list[dict[str, Any]]):
    header_row = detectar_header_requisicion(ws)
    cols = detectar_columnas_requisicion(ws, header_row)
    data_start = header_row + 2
    subtotal_row = find_totals_row(ws, data_start)
    filas_extra = preparar_filas_tabla(ws, data_start, subtotal_row, len(partidas))

    subtotal_row += filas_extra
    iva_row = subtotal_row + 1
    total_row = subtotal_row + 2

    # Limpiar rango de partidas, sin tocar la estructura.
    for row in range(data_start, data_start + len(partidas)):
        for col in [cols["par"], 2, cols["cantidad"], cols["unidad"], cols["descripcion"], cols["codigo"], cols["pu"], cols["total"]]:
            if col <= ws.max_column:
                set_cell(ws, ws.cell(row, col).coordinate, None)

    # Escribir partidas.
    for idx, partida in enumerate(partidas, start=1):
        row = data_start + idx - 1
        set_cell(ws, ws.cell(row, cols["par"]).coordinate, idx)
        set_cell(ws, ws.cell(row, cols["cantidad"]).coordinate, partida.get("cantidad"))
        set_cell(ws, ws.cell(row, cols["unidad"]).coordinate, partida.get("unidad"))
        set_cell(ws, ws.cell(row, cols["descripcion"]).coordinate, partida.get("descripcion"))
        set_cell(ws, ws.cell(row, cols["codigo"]).coordinate, partida.get("codigo"))
        set_cell(ws, ws.cell(row, cols["pu"]).coordinate, partida.get("precio_unitario"))
        set_cell(ws, ws.cell(row, cols["total"]).coordinate, f"={ws.cell(row, cols['cantidad']).coordinate}*{ws.cell(row, cols['pu']).coordinate}")

        # Forzar formato estable aunque la plantilla inserte filas extra.
        # Evita que los últimos productos salgan centrados/desacomodados.
        for col in [cols["par"], cols["cantidad"], cols["unidad"], cols["codigo"]]:
            cell = ws.cell(row, col)
            if not isinstance(cell, MergedCell):
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        desc_cell = ws.cell(row, cols["descripcion"])
        if not isinstance(desc_cell, MergedCell):
            desc_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        for col in [cols["pu"], cols["total"]]:
            cell = ws.cell(row, col)
            if not isinstance(cell, MergedCell):
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                cell.number_format = '"$"#,##0.00'

    # Totales.
    set_cell(ws, ws.cell(subtotal_row, max(1, cols["total"] - 1)).coordinate, "SUBTOTAL")
    set_cell(ws, ws.cell(subtotal_row, cols["total"]).coordinate, f"=SUM({ws.cell(data_start, cols['total']).coordinate}:{ws.cell(data_start + len(partidas) - 1, cols['total']).coordinate})")

    set_cell(ws, ws.cell(iva_row, max(1, cols["total"] - 1)).coordinate, "IVA")
    set_cell(ws, ws.cell(iva_row, cols["total"]).coordinate, f"={ws.cell(subtotal_row, cols['total']).coordinate}*0.16")

    set_cell(ws, ws.cell(total_row, max(1, cols["total"] - 1)).coordinate, "TOTAL")
    set_cell(ws, ws.cell(total_row, cols["total"]).coordinate, f"={ws.cell(subtotal_row, cols['total']).coordinate}+{ws.cell(iva_row, cols['total']).coordinate}")

    for row in range(data_start, total_row + 1):
        for col in [cols["pu"], cols["total"]]:
            ws.cell(row, col).number_format = '"$"#,##0.00'

    marcar_cierre_tabla(
        ws,
        primera_fila_vacia=data_start + len(partidas),
        ultima_fila_vacia=subtotal_row - 1,
        col_inicio=cols["par"],
        col_fin=cols["total"],
    )

    aplicar_campos_requisicion(ws, campos)


def detectar_header_oc(ws):
    for row in range(1, min(ws.max_row, 25) + 1):
        valores = [normalizar_mayus(get_cell(ws, ws.cell(row, col).coordinate)) for col in range(1, min(ws.max_column, 8) + 1)]
        joined = " ".join(valores)
        if "CANT" in joined and "DESCRIPCION" in joined and "IMPORTE" in joined:
            return row
    return 12


def llenar_orden_compra(ws, campos: dict[str, Any], partidas: list[dict[str, Any]]):
    header_row = detectar_header_oc(ws)
    data_start = header_row + 1

    subtotal = find_exact_cell(ws, "SUBTOTAL", start_row=data_start) or find_cell(ws, "SUBTOTAL", start_row=data_start)
    subtotal_row = subtotal[0] if subtotal else data_start + 20

    filas_extra = preparar_filas_tabla(ws, data_start, subtotal_row, len(partidas))
    subtotal_row += filas_extra

    for row in range(data_start, data_start + len(partidas)):
        for col in range(1, 6):
            set_cell(ws, ws.cell(row, col).coordinate, None)

    for idx, partida in enumerate(partidas, start=0):
        row = data_start + idx
        set_cell(ws, f"A{row}", partida.get("cantidad"))
        set_cell(ws, f"B{row}", partida.get("unidad"))
        set_cell(ws, f"C{row}", partida.get("descripcion"))
        set_cell(ws, f"D{row}", partida.get("precio_unitario"))
        set_cell(ws, f"E{row}", f"=A{row}*D{row}")

        for col in [1, 2]:
            cell = ws.cell(row, col)
            if not isinstance(cell, MergedCell):
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        desc_cell = ws.cell(row, 3)
        if not isinstance(desc_cell, MergedCell):
            desc_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        for col in [4, 5]:
            cell = ws.cell(row, col)
            if not isinstance(cell, MergedCell):
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                cell.number_format = '"$"#,##0.00'

    set_cell(ws, f"D{subtotal_row}", "SUBTOTAL")
    set_cell(ws, f"E{subtotal_row}", f"=SUM(E{data_start}:E{data_start + len(partidas) - 1})")

    # Buscar filas de FLETE, IVA, TOTAL.
    for label in ["FLETE", "IVA", "TOTAL USD", "TOTAL M.N"]:
        found = find_cell(ws, label, start_row=subtotal_row)
        if found:
            row, col, _ = found
            if "IVA" in label:
                set_cell(ws, ws.cell(row, col + 1).coordinate, f"=E{subtotal_row}*0.16")
            elif "TOTAL" in label:
                iva_found = find_cell(ws, "IVA", start_row=subtotal_row)
                iva_row = iva_found[0] if iva_found else subtotal_row + 2
                set_cell(ws, ws.cell(row, col + 1).coordinate, f"=E{subtotal_row}+E{iva_row}")

    # Campos principales.
    found = find_cell(ws, "ORDEN DE COMPRA")
    if found:
        row, col, _ = found
        set_cell(ws, ws.cell(row + 1, col).coordinate, campos.get("folio", ""))

    write_right_of_label(ws, "FECHA", campos.get("fecha", ""), 1)
    write_right_of_label(ws, "PROVEEDOR", campos.get("proveedor", ""), 1)
    write_right_of_label(ws, "FACTURAR A", campos.get("facturar_a", ""), 1)
    write_right_of_label(ws, "SOLICITA", campos.get("solicita", ""), 1)
    write_right_of_label(ws, "USO", campos.get("uso", ""), 1)
    write_right_of_label(ws, "ENTREGAR EN", campos.get("entregar_en", ""), 1)

    marcar_cierre_tabla(
        ws,
        primera_fila_vacia=data_start + len(partidas),
        ultima_fila_vacia=subtotal_row - 1,
        col_inicio=1,
        col_fin=5,
    )

    for row in range(data_start, subtotal_row + 8):
        for col in [4, 5]:
            ws.cell(row, col).number_format = '"$"#,##0.00'


def calcular_capacidad_requisicion(ws) -> int:
    header_row = detectar_header_requisicion(ws)
    data_start = header_row + 2
    subtotal_row = find_totals_row(ws, data_start)
    return max(1, subtotal_row - data_start)


def calcular_capacidad_orden_compra(ws) -> int:
    header_row = detectar_header_oc(ws)
    data_start = header_row + 1
    subtotal = find_exact_cell(ws, "SUBTOTAL", start_row=data_start) or find_cell(ws, "SUBTOTAL", start_row=data_start)
    subtotal_row = subtotal[0] if subtotal else data_start + 20
    return max(1, subtotal_row - data_start)


def dividir_partidas(partidas: list[dict[str, Any]], capacidad: int) -> list[list[dict[str, Any]]]:
    capacidad = max(1, capacidad)
    return [partidas[i:i + capacidad] for i in range(0, len(partidas), capacidad)] or [[]]


def titulo_hoja_pagina(nombre_base: str, pagina: int, total_paginas: int) -> str:
    base = nombre_hoja_valido(nombre_base or "Requisicion")
    if total_paginas <= 1:
        return base
    sufijo = f"-{pagina}"
    return nombre_hoja_valido(base[:31 - len(sufijo)] + sufijo)


def aplicar_cierres_visuales_a_xlsx(excel_bytes: bytes, wb, hojas_generadas: list[str]) -> bytes:
    resultado = excel_bytes
    for hoja in hojas_generadas:
        if hoja in wb.sheetnames:
            cierre_visual = getattr(wb[hoja], "_cierre_visual", None)
            if cierre_visual:
                resultado = agregar_cierre_diagonal_xlsx(resultado, hoja, cierre_visual)
    return resultado


def ocultar_bloque_inferior_requisicion_intermedia(ws):
    """
    En requisiciones paginadas, las hojas intermedias no deben mostrar
    observaciones, facturar a, recibe, entregar en ni firmas.
    Solo la última hoja conserva ese bloque inferior.
    """
    try:
        header_row = detectar_header_requisicion(ws)
        data_start = header_row + 2
        subtotal_row = find_totals_row(ws, data_start)
        total_row = subtotal_row + 2

        # Desde después del TOTAL hacia abajo está el bloque inferior:
        # CARGAR A, ID FAMILIA, OBSERVACIONES, FACTURAR A, RECIBE, FIRMAS, etc.
        inicio_ocultar = total_row + 1

        for row in range(inicio_ocultar, ws.max_row + 1):
            ws.row_dimensions[row].hidden = True
            ws.row_dimensions[row].height = 0

            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row, col)
                if isinstance(cell, MergedCell):
                    continue
                cell.value = None
                cell.border = Border()
                cell.fill = PatternFill(fill_type=None)
    except Exception:
        # Si no logra detectar el bloque, no detiene la generación.
        pass


def ocultar_bloque_inferior_oc_intermedia(ws):
    """
    Para órdenes de compra paginadas: dejar el bloque inferior solo en la última hoja.
    """
    try:
        header_row = detectar_header_oc(ws)
        data_start = header_row + 1
        subtotal = find_exact_cell(ws, "SUBTOTAL", start_row=data_start) or find_cell(ws, "SUBTOTAL", start_row=data_start)
        subtotal_row = subtotal[0] if subtotal else data_start + 20
        inicio_ocultar = subtotal_row + 3

        for row in range(inicio_ocultar, ws.max_row + 1):
            ws.row_dimensions[row].hidden = True
            ws.row_dimensions[row].height = 0

            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row, col)
                if isinstance(cell, MergedCell):
                    continue
                cell.value = None
                cell.border = Border()
                cell.fill = PatternFill(fill_type=None)
    except Exception:
        pass


def llenar_plantilla(
    plantilla_bytes: bytes,
    hoja_base: str,
    nombre_hoja_salida: str,
    campos: dict[str, Any],
    partidas: list[dict[str, Any]],
) -> BytesIO:
    wb = load_workbook(BytesIO(plantilla_bytes))
    ws_base = wb[hoja_base]

    tipo = detectar_tipo_plantilla(ws_base)
    nombre_salida_base = nombre_hoja_valido(nombre_hoja_salida)

    # En formatos fijos de requisición, si hay más partidas que filas disponibles,
    # NO se debe forzar todo en una sola hoja porque se mueve la parte inferior.
    # Se pagina automáticamente en hojas generadas: JAG25-1, JAG25-2, etc.
    if tipo == "orden_compra":
        capacidad_real = calcular_capacidad_orden_compra(ws_base)
    else:
        capacidad_real = calcular_capacidad_requisicion(ws_base)

    # Reservar una fila para el cierre visual cuando sea posible.
    capacidad_por_pagina = max(1, capacidad_real - 1)

    if len(partidas) > capacidad_por_pagina:
        grupos = dividir_partidas(partidas, capacidad_por_pagina)
    else:
        grupos = [partidas]

    total_paginas = len(grupos)

    # Crear todas las hojas generadas desde la plantilla limpia antes de llenar datos.
    hojas = [ws_base]
    for _ in range(1, total_paginas):
        hojas.append(wb.copy_worksheet(ws_base))

    hojas_generadas = []

    for idx, (ws, grupo) in enumerate(zip(hojas, grupos), start=1):
        titulo = titulo_hoja_pagina(nombre_salida_base, idx, total_paginas)

        # Evitar colisiones de nombres.
        nombre_final = titulo
        contador = 2
        while nombre_final in hojas_generadas:
            nombre_final = nombre_hoja_valido(f"{titulo[:27]}_{contador}")
            contador += 1

        ws.title = nombre_final
        hojas_generadas.append(ws.title)

        campos_pagina = dict(campos)

        if total_paginas > 1:
            # Mantener el folio original visible en la requisición.
            # El número de página va solo en el nombre de la hoja para no alterar el formato.
            campos_pagina["folio"] = campos.get("folio", "")

        if tipo == "orden_compra":
            llenar_orden_compra(ws, campos_pagina, grupo)
            if total_paginas > 1 and idx < total_paginas:
                ocultar_bloque_inferior_oc_intermedia(ws)
        else:
            llenar_requisicion(ws, campos_pagina, grupo)
            if total_paginas > 1 and idx < total_paginas:
                ocultar_bloque_inferior_requisicion_intermedia(ws)

    # Entrega limpia: dejar solo hojas generadas, no plantillas vacías.
    for nombre in list(wb.sheetnames):
        if nombre not in hojas_generadas:
            wb.remove(wb[nombre])

    wb.active = 0

    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    salida = BytesIO()
    wb.save(salida)
    excel_bytes = salida.getvalue()

    excel_bytes = aplicar_cierres_visuales_a_xlsx(excel_bytes, wb, hojas_generadas)

    resultado = BytesIO(excel_bytes)
    resultado.seek(0)
    return resultado


# =========================
# CONVERSIÓN A PDF
# =========================
def convertir_excel_a_pdf(excel_bytes: bytes, nombre_base: str) -> BytesIO:
    """
    Convierte el Excel generado a PDF usando LibreOffice en modo headless.
    En Streamlit Cloud se requiere packages.txt con libreoffice-calc.
    """
    ejecutable = shutil.which("libreoffice") or shutil.which("soffice")
    if not ejecutable:
        raise RuntimeError(
            "No se encontró LibreOffice en el servidor. "
            "Agrega un archivo packages.txt con libreoffice-calc y vuelve a desplegar."
        )

    nombre_seguro = re.sub(r"[^A-Za-z0-9_-]", "_", nombre_base).strip("_") or "REQUISICION"

    with tempfile.TemporaryDirectory() as tmp:
        carpeta = Path(tmp)
        entrada = carpeta / f"{nombre_seguro}.xlsx"
        salida_dir = carpeta / "pdf"
        salida_dir.mkdir(parents=True, exist_ok=True)

        entrada.write_bytes(excel_bytes)

        env = os.environ.copy()
        env["HOME"] = str(carpeta)

        comando = [
            ejecutable,
            "--headless",
            "--convert-to",
            "pdf",
            "--outdir",
            str(salida_dir),
            str(entrada),
        ]

        proceso = subprocess.run(
            comando,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            timeout=120,
            env=env,
        )

        pdf_esperado = salida_dir / f"{nombre_seguro}.pdf"
        if not pdf_esperado.exists():
            pdfs = list(salida_dir.glob("*.pdf"))
            if pdfs:
                pdf_esperado = pdfs[0]
            else:
                raise RuntimeError(
                    "LibreOffice no generó el PDF. "
                    f"Salida: {proceso.stdout} Error: {proceso.stderr}"
                )

        resultado = BytesIO(pdf_esperado.read_bytes())
        resultado.seek(0)
        return resultado


# =========================
# EXCEL SIMPLE
# =========================
def generar_excel_simple(campos: dict[str, Any], partidas: list[dict[str, Any]]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Requisicion"

    ws["A1"] = "REQUISICIÓN / COTIZACIÓN"
    ws["A1"].font = Font(bold=True, size=16)

    datos = [
        ("Folio", campos.get("folio", "")),
        ("Obra", campos.get("obra", "")),
        ("Proveedor", campos.get("proveedor", "")),
        ("Fecha", campos.get("fecha", "")),
        ("Semana", campos.get("semana", "")),
        ("No. Obra", campos.get("numero_obra", "")),
        ("Área", campos.get("area", "")),
        ("Partida", campos.get("partida", "")),
        ("Uso / notas", campos.get("uso", "")),
    ]

    row = 3
    for etiqueta, valor in datos:
        ws[f"A{row}"] = etiqueta
        ws[f"B{row}"] = valor
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

    header = row + 1
    encabezados = ["PAR", "Cantidad", "Unidad", "Descripción", "Código", "P.U.", "Importe"]

    fill_header = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, encabezado in enumerate(encabezados, start=1):
        c = ws.cell(header, col, encabezado)
        c.font = Font(bold=True)
        c.fill = fill_header
        c.border = border
        c.alignment = Alignment(horizontal="center")

    fila = header + 1
    for i, p in enumerate(partidas, start=1):
        values = [
            i,
            p.get("cantidad"),
            p.get("unidad"),
            p.get("descripcion"),
            p.get("codigo"),
            p.get("precio_unitario"),
            f"=B{fila}*F{fila}",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(fila, col, value)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        fila += 1

    subtotal = fila + 1
    ws.cell(subtotal, 6, "Subtotal")
    ws.cell(subtotal, 7, f"=SUM(G{header + 1}:G{fila - 1})")
    ws.cell(subtotal + 1, 6, "IVA 16%")
    ws.cell(subtotal + 1, 7, f"=G{subtotal}*0.16")
    ws.cell(subtotal + 2, 6, "Total")
    ws.cell(subtotal + 2, 7, f"=G{subtotal}+G{subtotal + 1}")

    for r in range(subtotal, subtotal + 3):
        ws.cell(r, 6).font = Font(bold=True)
        ws.cell(r, 7).font = Font(bold=True)
        ws.cell(r, 7).number_format = '"$"#,##0.00'

    for col, width in {"A": 8, "B": 12, "C": 12, "D": 55, "E": 16, "F": 14, "G": 16}.items():
        ws.column_dimensions[col].width = width

    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    salida = BytesIO()
    wb.save(salida)
    excel_bytes = salida.getvalue()

    cierre_visual = getattr(ws, "_cierre_visual", None)
    if cierre_visual:
        excel_bytes = agregar_cierre_diagonal_xlsx(excel_bytes, ws.title, cierre_visual)

    resultado = BytesIO(excel_bytes)
    resultado.seek(0)
    return resultado


# =========================
# EXTRACCIÓN DESDE EXCEL MATERIAL
# =========================
def detectar_folios_excel(ws_origen) -> list[str]:
    folios = []

    for fila in range(1, ws_origen.max_row + 1):
        texto = normalizar_mayus(ws_origen[f"A{fila}"].value)

        if re.match(r"^[A-Z]{1,12}-?\d{1,12}$", texto):
            if texto not in folios:
                folios.append(texto)

    return folios


def extraer_partidas_excel(origen_bytes: bytes, hoja_origen: str, folio: str):
    wb = load_workbook(BytesIO(origen_bytes), data_only=True)

    if hoja_origen not in wb.sheetnames:
        raise ValueError(f"No existe la hoja origen: {hoja_origen}")

    ws = wb[hoja_origen]
    folio_buscado = normalizar_mayus(folio)

    partidas = []

    for fila in range(1, ws.max_row + 1):
        folio_actual = normalizar_mayus(ws[f"A{fila}"].value)

        if folio_actual == folio_buscado:
            partidas.append(
                {
                    "fila_origen": fila,
                    "referencia": ws[f"A{fila}"].value,
                    "fecha": ws[f"B{fila}"].value,
                    "proveedor": ws[f"D{fila}"].value,
                    "codigo": ws[f"E{fila}"].value,
                    "cantidad": ws[f"F{fila}"].value,
                    "unidad": ws[f"G{fila}"].value,
                    "descripcion": ws[f"H{fila}"].value,
                    "precio_unitario": ws[f"O{fila}"].value,
                    "uso": ws[f"S{fila}"].value,
                }
            )

    if not partidas:
        raise ValueError(f"No encontré partidas con el folio/requisición: {folio}")

    primera = partidas[0]
    encabezado = {
        "folio": normalizar_mayus(folio),
        "fecha": primera.get("fecha"),
        "proveedor": primera.get("proveedor"),
        "uso": primera.get("uso"),
        "folios_detectados": [normalizar_mayus(folio)],
    }

    return encabezado, partidas


# =========================
# EXTRACCIÓN DESDE PDF
# =========================
def extraer_texto_pdf(pdf_bytes: bytes) -> str:
    if fitz is None:
        raise RuntimeError("Falta instalar PyMuPDF. Ejecuta: pip install pymupdf")

    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    return "\n".join([page.get_text("text") for page in doc])


def detectar_proveedor_pdf(texto: str) -> str:
    for linea in [l.strip() for l in texto.splitlines() if l.strip()]:
        if "GUTIERREZ FERRETEROS" in linea.upper():
            return linea
    return ""


def detectar_fecha_pdf(texto: str) -> str:
    meses = (
        "enero|febrero|marzo|abril|mayo|junio|julio|agosto|"
        "septiembre|setiembre|octubre|noviembre|diciembre"
    )
    fechas = re.findall(
        rf"\b\d{{1,2}}\s+(?:{meses}),?\s+\d{{4}}\b",
        texto,
        flags=re.IGNORECASE,
    )
    return fechas[0] if fechas else ""


def detectar_comentarios_pdf(texto: str) -> str:
    """
    Extrae el bloque de Comentarios del PDF, si existe.
    Ejemplo:
    PR-004//ALBAÑILERIA/CIMBRADO DE DADO/EXCAVACIONES...
    """
    m = re.search(
        r"Comentarios:\s*([\s\S]*?)(?:\n-\s*Cotización|\n-\s*Cotizacion|\nSubtotal:|\nTELEFONOS|$)",
        texto,
        flags=re.IGNORECASE,
    )

    if not m:
        return ""

    comentario = " ".join(m.group(1).split())

    # Limpieza de textos técnicos que no deben ir a observaciones.
    comentario = re.sub(r"Basado en operación de Retail One.*", "", comentario, flags=re.IGNORECASE).strip()
    comentario = re.sub(r"\b[A-Z0-9]{8,30}\b\s*$", "", comentario).strip()

    return comentario


def detectar_folios_pdf(texto: str) -> list[str]:
    lineas = [l.strip() for l in texto.splitlines() if l.strip()]
    folios = []

    def agregar(valor: str, al_inicio: bool = False):
        valor = normalizar_mayus(valor)
        valor = valor.replace("Ñ", "").replace("Ì", "").replace("Ó", "")

        if not valor or valor in {"GFE861208L17", "XAXX010101000"}:
            return

        if "/" in valor:
            return

        if valor not in folios:
            if al_inicio:
                folios.insert(0, valor)
            else:
                folios.append(valor)

    # 1. Folio de requisición dentro de comentarios: PR-004, VJ-295, etc.
    comentario = detectar_comentarios_pdf(texto)
    if comentario:
        m_req = re.search(r"\b[A-Z]{1,8}-\d{1,8}\b", comentario, flags=re.IGNORECASE)
        if m_req:
            agregar(m_req.group(0), al_inicio=True)

    # 2. Códigos de obra/cliente que vienen solos en una línea: JAG25, PARR001, LOTE108.
    for linea in lineas:
        candidato = normalizar_mayus(linea)
        if re.fullmatch(r"[A-Z]{2,12}\d{1,10}", candidato):
            agregar(candidato)

    # 3. Folio después de COTIZACIÓN/COTIZACION.
    for i, linea in enumerate(lineas):
        linea_mayus = normalizar_mayus(linea)
        if "COTIZACIÓN" in linea_mayus or "COTIZACION" in linea_mayus:
            for j in range(i + 1, min(i + 6, len(lineas))):
                candidato = normalizar_mayus(lineas[j])
                if re.fullmatch(r"[A-Z0-9]{6,30}", candidato):
                    agregar(candidato)
                    break

    # 4. Folio numérico superior.
    for linea in lineas[:10]:
        candidato = normalizar_mayus(linea)
        if re.fullmatch(r"\d{5,12}", candidato):
            agregar(candidato)
            break

    # 5. Requisiciones tipo VJ-295 en cualquier punto del texto.
    for m in re.findall(
        r"(?<![A-Z0-9/])VJ-\d{1,8}(?![A-Z0-9/])",
        texto,
        flags=re.IGNORECASE,
    ):
        agregar(m)

    return folios

def detectar_uso_pdf(texto: str, folio: str) -> str:
    if not folio:
        return ""

    folio_mayus = normalizar_mayus(folio)
    comentario = detectar_comentarios_pdf(texto)

    if comentario:
        # Si el comentario empieza con el folio, quitamos el folio y dejamos el resto.
        comentario_limpio = re.sub(
            rf"^\s*{re.escape(folio_mayus)}\s*[/\\|-]*\s*",
            "",
            comentario,
            flags=re.IGNORECASE,
        ).strip()

        # Si el folio seleccionado no era el del comentario, aun así dejamos comentario limpio.
        if comentario_limpio:
            return comentario_limpio

    # Para VJ, si no hay Comentarios, intenta leer notas.
    if re.fullmatch(r"VJ-\d{1,8}", folio_mayus):
        patron = re.compile(
            rf"({re.escape(folio_mayus)}[\s\S]*?)(?:\nNOTAS:|\nTELEFONOS|\nSubtotal:|\nComentarios:|$)",
            flags=re.IGNORECASE,
        )
        m = patron.search(texto)
        if m:
            return " ".join(m.group(1).split()).strip()

    return ""


CONTROL = {
    "IMPUESTOS", "TRASLADADOS", "BASE", "TIPO FACTOR", "IMPORTE", "TASA",
    "0.16", "H87", "KGM", "MTQ", "PIEZA", "KILO", "M3", "CÓDIGO", "CODIGO",
    "DESCRIPCIÓN", "DESCRIPCION", "CANTIDAD", "PRECIO", "COD ART. SAT",
}

UNIDADES = {
    "PIEZA": "PIEZA", "PZA": "PIEZA", "PZ": "PIEZA",
    "KILO": "KILO", "KG": "KILO", "KGM": "KILO",
    "M3": "M3", "MTQ": "M3",
    "METRO": "METRO", "MTS": "METRO", "LITRO": "LITRO",
    "LATA": "LATA", "CAJA": "CAJA", "ROLLO": "ROLLO",
}


def limpiar_descripcion(lineas: list[str], modo: str) -> str:
    partes = []

    for linea in lineas:
        valor = linea.strip()
        valor_mayus = valor.upper()

        if not valor:
            continue

        if valor_mayus.startswith("IMPUESTOS") or valor_mayus.startswith("TELEFONOS"):
            break

        if valor_mayus in CONTROL:
            continue

        if modo == "despues_de_precios":
            if re.fullmatch(r"\d{8}", valor):
                continue
            if es_dinero(valor):
                continue

        partes.append(valor)

    return " ".join(partes).strip()


def detectar_unidad(lineas: list[str], unidad_default: str) -> str:
    unidad = ""

    for linea in lineas:
        valor = linea.strip().upper()
        if valor in UNIDADES:
            unidad = UNIDADES[valor]

    return unidad or unidad_default


def es_candidato_modo_a(lineas: list[str], indice: int) -> bool:
    return (
        indice + 3 < len(lineas)
        and es_entero(lineas[indice])
        and es_numero(lineas[indice + 1])
        and es_dinero(lineas[indice + 2])
        and es_dinero(lineas[indice + 3])
    )


def buscar_inicio_modo_b(lineas: list[str], indice: int) -> int | None:
    limite = min(len(lineas) - 2, indice + 8)

    for j in range(indice + 2, limite):
        if es_numero(lineas[j]) and es_dinero(lineas[j + 1]) and es_dinero(lineas[j + 2]):
            return j

    return None


def extraer_partidas_pdf(pdf_bytes: bytes, unidad_default: str):
    texto = extraer_texto_pdf(pdf_bytes)
    lineas = [l.strip() for l in texto.splitlines() if l.strip()]

    folios = detectar_folios_pdf(texto)
    proveedor = detectar_proveedor_pdf(texto)
    fecha = detectar_fecha_pdf(texto)

    partidas = []
    i = 0

    while i < len(lineas):
        if not es_entero(lineas[i]):
            i += 1
            continue

        codigo = lineas[i]

        # Modo A: código / cantidad / precio / importe / descripción.
        if es_candidato_modo_a(lineas, i):
            cantidad = limpiar_numero(lineas[i + 1])
            numero_1 = limpiar_numero(lineas[i + 2])
            numero_2 = limpiar_numero(lineas[i + 3])
            precio, importe = elegir_precio_importe(cantidad, numero_1, numero_2)

            bloque = []
            j = i + 4
            while j < len(lineas):
                if es_candidato_modo_a(lineas, j):
                    break
                if es_entero(lineas[j]) and buscar_inicio_modo_b(lineas, j) is not None:
                    break
                bloque.append(lineas[j])
                j += 1

            descripcion_base = []
            for linea in bloque:
                if linea.upper().startswith("IMPUESTOS"):
                    break
                descripcion_base.append(linea)

            partidas.append(
                {
                    "codigo": codigo,
                    "cantidad": cantidad,
                    "unidad": detectar_unidad(bloque, unidad_default),
                    "descripcion": limpiar_descripcion(descripcion_base, "despues_de_precios"),
                    "precio_unitario": precio,
                    "importe": importe,
                }
            )
            i = j
            continue

        # Modo B: código / descripción / cantidad / precio-importe.
        inicio_numeros = buscar_inicio_modo_b(lineas, i)
        if inicio_numeros is not None:
            cantidad = limpiar_numero(lineas[inicio_numeros])
            numero_1 = limpiar_numero(lineas[inicio_numeros + 1])
            numero_2 = limpiar_numero(lineas[inicio_numeros + 2])
            precio, importe = elegir_precio_importe(cantidad, numero_1, numero_2)
            descripcion = limpiar_descripcion(lineas[i + 1:inicio_numeros], "antes_de_precios")

            if descripcion:
                partidas.append(
                    {
                        "codigo": codigo,
                        "cantidad": cantidad,
                        "unidad": unidad_default,
                        "descripcion": descripcion,
                        "precio_unitario": precio,
                        "importe": importe,
                    }
                )

            i = inicio_numeros + 3
            continue

        i += 1

    if not partidas:
        raise ValueError(
            "No pude extraer partidas del PDF. Puede ser escaneado como imagen "
            "o tener un formato todavía no registrado."
        )

    folio_sugerido = folios[0] if folios else ""
    encabezado = {
        "folio": folio_sugerido,
        "fecha": fecha,
        "proveedor": proveedor,
        "uso": detectar_uso_pdf(texto, folio_sugerido),
        "folios_detectados": folios,
    }

    return encabezado, partidas, texto



def sugerir_campos_desde_comentarios(texto: str, folio: str) -> dict[str, str]:
    """
    Convierte comentarios tipo:
    PR-004//ALBAÑILERIA/CIMBRADO DE DADO/EXCAVACIONES Y PERFILADOS...
    en sugerencias editables.
    """
    comentario = detectar_comentarios_pdf(texto)
    if not comentario:
        return {}

    limpio = comentario
    if folio:
        limpio = re.sub(
            rf"^\s*{re.escape(normalizar_mayus(folio))}\s*[/\\|-]*\s*",
            "",
            limpio,
            flags=re.IGNORECASE,
        ).strip()

    partes = [p.strip() for p in re.split(r"/+", limpio) if p.strip()]
    sugerencias = {"uso": limpio}

    if partes:
        sugerencias["area"] = partes[0]

    if len(partes) >= 2:
        # Usa la última sección como partida, pero si contiene EXCAVACIONES, deja esa palabra como partida.
        ultima = partes[-1]
        if "EXCAVACION" in normalizar_mayus(ultima) or "EXCAVACIONES" in normalizar_mayus(ultima):
            sugerencias["partida"] = "EXCAVACIONES"
        else:
            sugerencias["partida"] = ultima

    return sugerencias


# =========================
# INTERFAZ
# =========================
st.sidebar.title("Uso rápido")
st.sidebar.write("1. Elige Excel o PDF.")
st.sidebar.write("2. Sube el archivo origen.")
st.sidebar.write("3. Selecciona una plantilla integrada o sube una propia.")
st.sidebar.write("4. Edita datos generales.")
st.sidebar.write("5. Genera y descarga el Excel o PDF con cierre visual.")
st.sidebar.divider()
st.sidebar.info("La app incluye FORMATOS.xlsx como banco de plantillas.")

st.title("Sistema de automatización de requisiciones")

st.write(
    "Sistema web para generar requisiciones automáticamente a partir de cotizaciones PDF o archivos Excel. "
    "Selecciona el formato requerido, revisa los datos detectados y descarga el archivo listo para usar."
)

modo = st.radio(
    "Tipo de archivo origen",
    ["Excel", "PDF"],
    horizontal=True,
)

st.divider()

col1, col2 = st.columns(2)

with col1:
    if modo == "Excel":
        archivo_origen = st.file_uploader("1. Archivo de origen", type=["xlsx", "xlsm"])
    else:
        archivo_origen = st.file_uploader("1. Archivo de origen", type=["pdf"])

with col2:
    usar_integrada = st.checkbox(
        "Usar plantillas integradas FORMATOS.xlsx",
        value=BUILTIN_TEMPLATE_PATH.exists(),
        disabled=not BUILTIN_TEMPLATE_PATH.exists(),
    )

    archivo_plantilla = None
    if not usar_integrada:
        archivo_plantilla = st.file_uploader("2. Subir plantilla Excel", type=["xlsx", "xlsm"])

datos_encabezado = None
partidas = []
texto_pdf = ""
folio_final = ""

if archivo_origen and modo == "Excel":
    origen_bytes = archivo_origen.getvalue()
    wb_tmp = load_workbook(BytesIO(origen_bytes), data_only=True)
    hojas = wb_tmp.sheetnames
    index_default = hojas.index("MATERIAL VJ") if "MATERIAL VJ" in hojas else 0
    hoja_origen = st.selectbox("Hoja origen", hojas, index=index_default)

    folios_excel = detectar_folios_excel(wb_tmp[hoja_origen])
    if folios_excel:
        folio_final = st.selectbox("Folio/requisición a buscar", folios_excel)
    else:
        folio_final = st.text_input("Folio/requisición a buscar", value="")

    if folio_final:
        try:
            datos_encabezado, partidas = extraer_partidas_excel(origen_bytes, hoja_origen, folio_final)
        except Exception as e:
            st.warning(str(e))

elif archivo_origen and modo == "PDF":
    origen_bytes = archivo_origen.getvalue()
    unidad_default = st.text_input("Unidad por defecto si el PDF no trae unidad", value="PIEZA")

    try:
        datos_encabezado, partidas, texto_pdf = extraer_partidas_pdf(origen_bytes, unidad_default)
        folios_detectados = datos_encabezado.get("folios_detectados", [])

        if folios_detectados:
            opcion = st.selectbox("Folio/requisición detectado", folios_detectados + ["ESCRIBIR OTRO"])
            if opcion == "ESCRIBIR OTRO":
                folio_final = st.text_input("Escribe el folio/requisición", value="")
            else:
                folio_final = opcion
        else:
            folio_final = st.text_input("Folio/requisición", value="")

        datos_encabezado["folio"] = folio_final
        datos_encabezado["uso"] = detectar_uso_pdf(texto_pdf, folio_final)
        datos_encabezado.update(sugerir_campos_desde_comentarios(texto_pdf, folio_final))

        st.success(f"PDF leído. Partidas detectadas: {len(partidas)}")
        with st.expander("Ver datos detectados"):
            st.write(f"Folio/requisición: {folio_final}")
            st.write(f"Fecha: {datos_encabezado.get('fecha')}")
            st.write(f"Proveedor: {datos_encabezado.get('proveedor')}")
            st.write(f"Folios detectados: {', '.join(folios_detectados) if folios_detectados else 'Sin folios detectados'}")
        with st.expander("Ver texto extraído del PDF para diagnóstico"):
            st.text(texto_pdf[:8000])
    except Exception as e:
        st.error("No pude leer correctamente el PDF.")
        st.exception(e)

# Plantillas
plantilla_bytes = None
hoja_base = ""
campos_base = {}
tipo_plantilla = ""

if usar_integrada and BUILTIN_TEMPLATE_PATH.exists():
    plantilla_bytes = BUILTIN_TEMPLATE_PATH.read_bytes()
elif archivo_plantilla:
    plantilla_bytes = archivo_plantilla.getvalue()

if plantilla_bytes:
    wb_plantilla_tmp = load_workbook(BytesIO(plantilla_bytes))
    hojas_plantilla = wb_plantilla_tmp.sheetnames
    indice_formato = 0
    folio_para_formato = normalizar_mayus(folio_final)

    formato_sugerido = ""
    m_formato = re.match(r"^([A-Z]+)-", folio_para_formato)
    if m_formato:
        formato_sugerido = m_formato.group(1)

    if formato_sugerido in hojas_plantilla:
        indice_formato = hojas_plantilla.index(formato_sugerido)

    hoja_base = st.selectbox("Formato / hoja de plantilla", hojas_plantilla, index=indice_formato)
    ws_base = wb_plantilla_tmp[hoja_base]
    tipo_plantilla = detectar_tipo_plantilla(ws_base)

    if tipo_plantilla == "orden_compra":
        campos_base = leer_campos_orden_compra(ws_base)
        st.info("Plantilla tipo Orden de Compra detectada.")
    else:
        campos_base = leer_campos_requisicion(ws_base)
        st.info("Plantilla tipo Requisición detectada. Se conservarán logos, formato, celdas combinadas, observaciones y firmas.")

        if partidas:
            capacidad_tmp = max(1, calcular_capacidad_requisicion(ws_base) - 1)
            if len(partidas) > capacidad_tmp:
                paginas_tmp = (len(partidas) + capacidad_tmp - 1) // capacidad_tmp
                st.warning(
                    f"Esta requisición tiene {len(partidas)} partidas y el formato solo permite "
                    f"{capacidad_tmp} por hoja sin mover la parte inferior. "
                    f"Se generarán {paginas_tmp} hojas/páginas automáticamente."
                )
else:
    st.warning("No hay plantilla seleccionada. Se generará un Excel simple sin logos ni formato de requisición.")

st.divider()

if datos_encabezado is None:
    datos_encabezado = {}

if folio_final:
    campos_base["folio"] = folio_final

if datos_encabezado.get("proveedor"):
    campos_base["proveedor"] = datos_encabezado.get("proveedor")

if datos_encabezado.get("fecha"):
    campos_base["fecha"] = datos_encabezado.get("fecha")

if datos_encabezado.get("uso"):
    campos_base["uso"] = datos_encabezado.get("uso")

if datos_encabezado.get("area"):
    campos_base["area"] = datos_encabezado.get("area")

if datos_encabezado.get("partida"):
    campos_base["partida"] = datos_encabezado.get("partida")

st.subheader("Datos editables del formato")

col_a, col_b, col_c = st.columns(3)

with col_a:
    campo_folio = st.text_input("No. / Folio", value=str(campos_base.get("folio", "")))
    campo_obra = st.text_input("Obra", value=str(campos_base.get("obra", "")))
    campo_proveedor = st.text_input("Proveedor", value=str(campos_base.get("proveedor", "")))
    campo_fecha = st.text_input("Fecha", value=str(campos_base.get("fecha", "")))

with col_b:
    campo_semana = st.text_input("Semana", value=str(campos_base.get("semana", "")))
    campo_numero_obra = st.text_input("No. Obra", value=str(campos_base.get("numero_obra", "")))
    campo_id_familia = st.text_input("ID Familia / Cargar a", value=str(campos_base.get("id_familia", "")))
    campo_etapa = st.text_input("Etapa", value=str(campos_base.get("etapa", "")))

with col_c:
    campo_area = st.text_input("Área", value=str(campos_base.get("area", "")))
    campo_partida = st.text_input("Partida", value=str(campos_base.get("partida", "")))
    campo_facturar_a = st.text_input("Facturar a", value=str(campos_base.get("facturar_a", "")))

campo_uso = st.text_area("Uso / observaciones", value=str(campos_base.get("uso", "")), height=100)

col_d, col_e = st.columns(2)

with col_d:
    campo_solicita = st.text_input("Solicita", value=str(campos_base.get("solicita", "")))
    campo_recibe = st.text_input("Recibe", value=str(campos_base.get("recibe", "")))
    campo_entregar_en = st.text_area("Entregar en", value=str(campos_base.get("entregar_en", "")), height=70)

with col_e:
    campo_firma_izquierda = st.text_input("Firma izquierda", value=str(campos_base.get("firma_izquierda", "")))
    campo_puesto_izquierdo = st.text_input("Puesto firma izquierda", value=str(campos_base.get("puesto_izquierdo", "")))
    campo_firma_derecha = st.text_input("Firma derecha", value=str(campos_base.get("firma_derecha", "")))
    campo_puesto_derecho = st.text_input("Puesto firma derecha", value=str(campos_base.get("puesto_derecho", "")))

nombre_hoja_salida = st.text_input(
    "Nombre de la hoja en el archivo generado",
    value=nombre_hoja_valido(campo_folio or hoja_base or "Requisicion"),
)

campos_finales = {
    "folio": campo_folio,
    "obra": campo_obra,
    "proveedor": campo_proveedor,
    "fecha": campo_fecha,
    "semana": campo_semana,
    "numero_obra": campo_numero_obra,
    "id_familia": campo_id_familia,
    "etapa": campo_etapa,
    "area": campo_area,
    "partida": campo_partida,
    "uso": campo_uso,
    "solicita": campo_solicita,
    "recibe": campo_recibe,
    "entregar_en": campo_entregar_en,
    "facturar_a": campo_facturar_a,
    "firma_izquierda": campo_firma_izquierda,
    "firma_derecha": campo_firma_derecha,
    "puesto_izquierdo": campo_puesto_izquierdo,
    "puesto_derecho": campo_puesto_derecho,
}

st.divider()

if partidas:
    st.subheader("Vista previa de partidas")
    vista = []
    for p in partidas:
        vista.append(
            {
                "Código": p.get("codigo"),
                "Cantidad": p.get("cantidad"),
                "Unidad": p.get("unidad"),
                "Descripción": p.get("descripcion"),
                "Precio unitario": p.get("precio_unitario"),
                "Importe detectado": p.get("importe"),
            }
        )
    st.dataframe(vista, use_container_width=True)

if st.button("Generar archivo", type="primary"):
    if not archivo_origen:
        st.error("Sube primero el archivo origen.")
    elif not partidas:
        st.error("No hay partidas detectadas.")
    elif not campo_folio:
        st.error("Selecciona o escribe el No. / Folio.")
    else:
        try:
            if plantilla_bytes and hoja_base:
                archivo_generado = llenar_plantilla(
                    plantilla_bytes=plantilla_bytes,
                    hoja_base=hoja_base,
                    nombre_hoja_salida=nombre_hoja_salida,
                    campos=campos_finales,
                    partidas=partidas,
                )
            else:
                archivo_generado = generar_excel_simple(campos_finales, partidas)

            nombre = re.sub(r"[^A-Z0-9_-]", "_", normalizar_mayus(campo_folio)) or "REQUISICION"
            excel_bytes_final = archivo_generado.getvalue()

            st.success("Archivo generado correctamente.")

            col_excel, col_pdf = st.columns(2)

            with col_excel:
                st.download_button(
                    "Descargar Excel generado",
                    data=excel_bytes_final,
                    file_name=f"{nombre}_automatico.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            with col_pdf:
                try:
                    pdf_generado = convertir_excel_a_pdf(excel_bytes_final, nombre)
                    st.download_button(
                        "Descargar PDF generado",
                        data=pdf_generado.getvalue(),
                        file_name=f"{nombre}_automatico.pdf",
                        mime="application/pdf",
                    )
                except Exception as pdf_error:
                    st.warning(
                        "Se generó el Excel, pero no se pudo crear el PDF en este servidor. "
                        "Verifica que packages.txt tenga libreoffice-calc y vuelve a desplegar."
                    )
                    with st.expander("Ver detalle del error PDF"):
                        st.write(str(pdf_error))
        except Exception as e:
            st.error("No se pudo generar el archivo.")
            st.exception(e)
